import os
import re
import pymupdf
from langchain.docstore.document import Document
from langchain_community.document_loaders import WebBaseLoader
# from llama_parse import LlamaParse

os.getenv("USER_AGENT")

def load_documents(file_type, file, filename):
    if file_type == FileType.PDF:
        documents = load_pdf(file, filename)
    elif file_type == FileType.WORD:
        documents = load_word(file, filename)
    elif file_type == FileType.EXCEL:
        documents = load_excel(file, filename)

    return documents


def load_pdf(file, filename, page_overlap=256):
    documents = []
    try:
        # Open the PDF file
        doc = pymupdf.open(file, filetype="pdf")

        # Extract text from each page
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            text = page.get_text("text")
            text = clean_extra_whitespace(text)
            text = group_broken_paragraphs(text)

            # Obtener el overlap de la página anterior.
            if page_num > 0:
                prev_page = doc.load_page(page_num - 1)
                prev_page_text = prev_page.get_text("text")
                prev_page_text = clean_extra_whitespace(prev_page_text)
                prev_page_text = group_broken_paragraphs(prev_page_text)
                text = prev_page_text[-page_overlap:] + " " + text

            # Obtener el overlap de la página siguiente.
            if page_num < len(doc) - 1:
                next_page = doc.load_page(page_num + 1)
                next_page_text = next_page.get_text("text")
                next_page_text = clean_extra_whitespace(next_page_text)
                next_page_text = group_broken_paragraphs(next_page_text)
                text += " " + next_page_text[:page_overlap]

            metadata = {
                "source": filename,
                "file_type": FileType.PDF,
                "page_number": page_num + 1,
                "sheet_name": "",
            }

            document = Document(page_content=text, metadata=metadata)
            documents.append(document)
    except Exception as e:
        print(e)
        raise

    return documents


def load_webpage(url):
    documents = []
    try:
        loader = WebBaseLoader(url)
        docs = loader.load()
        for doc in docs:
            text = doc.page_content
            text = clean_extra_whitespace(text)
            text = group_broken_paragraphs(text)

            metadata = {
                "source": url,
                "file_type": FileType.WEBPAGE,
                "page_number": -1,
                "sheet_name": "",
            }

            document = Document(page_content=text, metadata=metadata)
            documents.append(document)
    except Exception as e:
        print(f"An error occurred while loading {e}: {e}")
        raise

    return documents


def load_word(file, filename):
    documents = []
    try:
        parser = LlamaParse(result_type="markdown")
        docs = parser.aload_data(file, extra_info={"file_name": filename})
        for page_num, doc in enumerate(docs):
            text = doc.text
            metadata = {
                "source": filename,
                "file_type": FileType.WORD,
                "page_number": page_num + 1,
                "sheet_name": "",
            }

            document = Document(page_content=text, metadata=metadata)
            documents.append(document)
    except Exception as e:
        print(e)
        raise

    return documents


import docx2txt
import re
from typing import List
from langchain.schema import Document

def split_into_pages(text: str) -> List[str]:
    """
    Divide el texto en páginas basándose en saltos de página.
    """
    return text.split('\f')

def load_docx(file_path: str, filename: str) -> List[Document]:
    """
    Carga texto de un archivo DOCX, lo divide en páginas y crea objetos Document con overlap.
    """
    # Extraer texto del archivo DOCX
    text = docx2txt.process(file_path)
    
    # Limpiar y agrupar párrafos
    text = clean_extra_whitespace(text)
    text = group_broken_paragraphs(text)
    
    # Dividir en páginas
    pages = split_into_pages(text)
    
    documents = []
    overlap_words = 8

    for i, page_content in enumerate(pages):
        # Preparar el contenido con overlap
        content_parts = []
        
        # Agregar overlap antes
        if i > 0:
            prev_page_words = pages[i-1].split()
            content_parts.append(" ".join(prev_page_words[-overlap_words:]))
        
        # Agregar contenido de la página actual
        content_parts.append(page_content)
        
        # Agregar overlap después
        if i < len(pages) - 1:
            next_page_words = pages[i+1].split()
            content_parts.append(" ".join(next_page_words[:overlap_words]))
        
        # Crear objeto Document
        doc = Document(
            page_content=" ".join(content_parts),
            metadata = {
                "source": filename,
                "file_type": FileType.WORD,
                "page_number": i + 1,
                "sheet_name": "",
            }
        )
        documents.append(doc)
    
    # print("------------------------------------------------------------------------------>>>\nDocx documents:")
    # print(documents)
    
    return documents


def load_excel(file, filename):
    documents = []
    try:
        parser = LlamaParse(result_type="markdown")
        docs = parser.aload_data(file, extra_info={"file_name": filename})
        for doc in docs:
            text = doc.text
            metadata = {
                "source": filename,
                "file_type": FileType.EXCEL,
                "page_number": -1,
                "sheet_name": extract_sheet_name(text),
            }

            document = Document(page_content=text, metadata=metadata)
            documents.append(document)
    except Exception as e:
        print(e)
        raise

    return documents


import openpyxl
def load_xlsx(file_path, filename):
    """
    Loads text from an XLSX file, including sheet names in metadata.
    """
    wb = openpyxl.load_workbook(file_path)
    documents = []
    for sheet_num, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        text = ""
        for row in ws.iter_rows(values_only=True):
            text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
        text = clean_extra_whitespace(text)
        text = group_broken_paragraphs(text)
        if text:  # Only add if the text is not empty
            document = Document(
                page_content=text,
                metadata = {
                    "source": filename,
                    "file_type": FileType.EXCEL,
                    "page_number": -1,
                    "sheet_name": sheet_name,
                }
            )
            documents.append(document)

    # Agregado
    # print ("------------------------------------------------------------------------------>>>\nXlsx documents:")
    # print (documents)

    return documents


def clean_extra_whitespace(text):
    """
    Cleans extra whitespace from the provided text.

    Parameters:
    - text: A string representing the text to be cleaned.

    Returns:
    - A string with extra whitespace removed.
    """
    return " ".join(text.split())


def group_broken_paragraphs(text):
    """
    Groups broken paragraphs in the provided text.

    Parameters:
    - text: A string representing the text to be processed.

    Returns:
    - A string with broken paragraphs grouped.
    """
    return text.replace("\n", " ").replace("\r", " ")


def extract_sheet_name(text):
    match = re.match(r"^# (.*?)\n", text)
    if match:
        return match.group(1)
    return None

# Tipos de archivo y extensiones permitidas
class FileType:
    PDF = "pdf"
    WORD = "word"
    EXCEL = "excel"
    WEBPAGE = "webpage"